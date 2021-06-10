
#  Python imports
import threading
import json, time
import xlsxwriter 
import pandas as pd
from numpy import nan
import openpyxl as op
from bs4 import BeautifulSoup
from selenium import webdriver 


book_prefix = {'United States' : 'https://www.amazon.com'}

class audible:
    count = 1
    sub_level = 0
    sub_names = {}
    categories = []
    book_number = 50
    audible_categories = {}
    audible_filename = 'audible_data.xlsx'
    data_fields =  ['category', 'subcat-1', 'subcat-2', 'subcat-3', 'subcat-4']

    def __init__(self):
        setting =  pd.read_excel('settings.xlsx', 'settings')
        datas = setting['audible-data-fields']
        cats  = setting['audible-categories']
        self.book_number = int(setting['book-number'][0])
        for i in datas.index:
            if datas[i] is not nan:    self.data_fields.append(datas[i])
        for i in cats.index:
            if cats[i] is not nan:     self.categories.append(cats[i])
        self.audible_categories = self.update_category_list('audible_list.json')

    def update_category_list(self, filename):
        with open(filename, 'r+') as read_file:
            return json.load(read_file)
    
    def scrape_category(self):
        for cat in self.categories:
            self.create_excel_file(cat, self.audible_filename)
            break
        for category in self.categories:   # Create new thread for category
            self.sub_level = 1
            self.sub_names = {'category': category, 'subcat-1' : 'null', 'subcat-2' : 'null', 'subcat-3' : 'null', 'subcat-4' : 'null'}
            workbook = op.load_workbook(self.audible_filename, False)
            try:    workbook[category]
            except:
                workbook.create_sheet(category)
                worksheet = workbook[category]
                worksheet.append(self.headers())
                workbook.save(self.audible_filename)
                workbook.close()
            try:
                subcategories = self.audible_categories[ category ]
                t = threading.Thread(target=self.helper_category_books, args=(subcategories, self.audible_filename, ))
                t.start()
                t.join()
                time.sleep(1)
            except: print ("Error: unable to start new thread")

    def update_subnames(self, subcat):
        print('\n', self.sub_names)
        if self.sub_level == 1: self.sub_names['subcat-1'] = subcat
        if self.sub_level == 2: self.sub_names['subcat-2'] = subcat
        if self.sub_level == 3: self.sub_names['subcat-3'] = subcat
        if self.sub_level == 4: self.sub_names['subcat-4'] = subcat   

    def helper_category_books(self, subcategories, filename):
        for subcat in subcategories:
            if not type(subcategories[subcat]) is dict:
                self.update_subnames(subcat)
                try:    
                    self.category_books(filename, subcategories[subcat])
                except: print ("Error: unable to start new thread")
            else:
                self.update_subnames(subcat)
                self.sub_level += 1
                self.helper_category_books(subcategories[subcat], filename)
        self.update_subnames('null')
        self.sub_level -= 1
            
    def category_books(self, filename, link):
        books = []  # for book-data
        browser = webdriver.Chrome('chromedriver.exe') 
        try:    browser.get(link)
        except: return print('Failed to Load')
        source = browser.page_source
        soup = BeautifulSoup(source, features='lxml')
        book_sections = soup.find_all('div', attrs={'class':'a-section a-spacing-none aok-relative'})
        book_count = 1
        for book in book_sections:
            book_count += 1
            a_tags = book.find_all('a', attrs={'class':'a-link-normal'})
            book_details_link = book_prefix['United States'] + a_tags[0]['href']
            try:
                browser.get(book_details_link)
                source = browser.page_source
                soup = BeautifulSoup(source, features='lxml')
                title = soup.find('span', attrs={'id':'productTitle'}).get_text().strip('\n')
                span = soup.find_all('span', attrs={'class', 'author notFaded'})
                author = span[0].find('a', attrs={'class':'a-link-normal'}).get_text()
                rating = soup.find('span', attrs={'id':'acrCustomerReviewText'}).get_text().replace(' ratings', '')
                stars = soup.find('span', attrs={'class': 'reviewCountTextLinkedHistogram noUnderline'})['title'].replace(' out of 5 stars', '')
                table = soup.find('table', attrs={'class':'a-keyvalue a-vertical-stripes a-span6'})
                table = table.find('tbody')
                tr_tags = table.find_all('tr')
                tr_list = []
                for i in range(len(tr_tags)-1):
                    tr_list.append(tr_tags[i])

                details = {}
                for name            in self.sub_names:      details[name]           = self.sub_names[name]
                if 'Title'          in self.data_fields:    details['Title']        = title
                if 'Web-Link'       in self.data_fields:    details['Web-Link']     = book_details_link
                if 'Author'         in self.data_fields:    details['Author']       = author
                if 'Ratings'        in self.data_fields:    details['Ratings']      = rating
                if 'Stars'          in self.data_fields:    details['Stars']        = stars

                for tr in tr_list:
                    span = tr.find('th')
                    span = span.find('span')
                    heading = span.get_text() 
                    data = tr.find('td')
                    try:    data = data.find('span').get_text()
                    except: data = data.find('a').get_text()
                    details[heading] = data

                bst_heading = table.find('th', attrs={'class':'a-color-secondary a-size-base prodDetSectionEntry'}).get_text().replace('\n','')
                bst_tr = None
                for tr in tr_tags:  bst_tr = tr
                td = bst_tr.find('td')
                span = td.find('span')
                spans = span.find_all('span')
                bst_data = []
                for span in spans:
                    try:    bst_data.append(span.get_text().split('(')[0])
                    except:    bst_data.append(span.get_text())
                details[bst_heading] = bst_data
                print( 'A-', self.count, '. ', details, '\n')
                self.count += 1
                books.append(details)
            except: continue
            if book_count > self.book_number:   break

        # Saving in Excel File  
        self.write_to_excel(filename, books)
        browser.close()

    def create_excel_file(self, category_name, filename):
        # creating new excle file
        workbook = xlsxwriter.Workbook(filename)
        workbook.add_worksheet(category_name)
        workbook.close()
        workbook = op.load_workbook(filename, False)
        worksheet = workbook[category_name]
        worksheet.append(self.headers())
        workbook.save(filename)
        workbook.close()

    def write_to_excel(self, filename, books=[]):
        workbook = op.load_workbook(filename, False)
        worksheet = workbook[self.sub_names['category']]
        for book in books:
            data = []
            for data_field in self.data_fields:
                try:
                    if data_field == 'Best Sellers Rank':
                        for v in book[data_field]: data.append(v)
                    else:   data.append(book[data_field])
                except: data.append('N/A')
            worksheet.append(data)
        workbook.save(filename)
        workbook.close()

    def headers(self):
        header = []
        for data_field in self.data_fields:
            header.append(data_field)
        return header


class kindle:
    count = 1
    sub_level = 0
    sub_names = {}
    categories = []
    book_number = 50
    kindle_categories = {}
    kindle_filename = 'kindle_data.xlsx'
    data_fields =  ['category', 'subcat-1', 'subcat-2', 'subcat-3', 'subcat-4']
    
    def __init__(self):
        setting =  pd.read_excel('settings.xlsx', 'settings')
        datas = setting['kindle-data-fields']
        cats  = setting['kindle-categories']
        self.book_number = int(setting['book-number'][0])
        for i in datas.index:
            if datas[i] is not nan:    self.data_fields.append(datas[i])
        for i in cats.index:
            if cats[i] is not nan:     self.categories.append(cats[i])
        self.kindle_categories = self.update_category_list('kindle_list.json')

    def update_category_list(self, filename):
        with open(filename, 'r+') as read_file:
            return json.load(read_file)

    def scrape_category(self):
        count = 1
        for cat in self.categories:
            self.create_excel_file(cat, self.kindle_filename)
            break
        for category in self.categories:   # Create new thread for category
            self.sub_level = 1
            self.sub_names = {'category': category, 'subcat-1' : 'null', 'subcat-2' : 'null', 'subcat-3' : 'null', 'subcat-4' : 'null'}
            workbook = op.load_workbook(self.kindle_filename, False)
            try:
                workbook[category]
            except:
                workbook.create_sheet(category)
                worksheet = workbook[category]
                worksheet.append(self.headers())
                workbook.save(self.kindle_filename)
                workbook.close()
            print(category)
            try:
                subcategories = self.kindle_categories[ category ]
                t = threading.Thread(target=self.helper_category_books, args=(subcategories, self.kindle_filename, ))
                t.start()
                t.join()
                time.sleep(1)
            except:
                print ("Error: unable to start new thread")
            count += 1
    
    def update_subnames(self, subcat):
        print('\n', self.sub_names)
        if self.sub_level == 1: self.sub_names['subcat-1'] = subcat
        if self.sub_level == 2: self.sub_names['subcat-2'] = subcat
        if self.sub_level == 3: self.sub_names['subcat-3'] = subcat
        if self.sub_level == 4: self.sub_names['subcat-4'] = subcat   

    def helper_category_books(self, subcategories, filename):
        for subcat in subcategories:
            if not type(subcategories[subcat]) is dict:
                self.update_subnames(subcat)
                try:   
                    self.category_books(filename, subcategories[subcat])
                except:
                    print ("Error: unable to start new thread")
            else:
                self.update_subnames(subcat)
                self.sub_level += 1
                self.helper_category_books(subcategories[subcat], filename)
        self.update_subnames('null')
        self.sub_level -= 1
    
    def category_books(self, filename, link):
        books = []  # for book-data
        browser = webdriver.Chrome('chromedriver.exe') 
        browser.set_window_position(680, 30)
        browser.set_window_size(700, 700)
        try:    browser.get(link)
        except: return print('Failed to Load')
        source = browser.page_source
        soup = BeautifulSoup(source, features='lxml')
        book_sections = soup.find_all('div', attrs={'class':'a-section a-spacing-none aok-relative'})
        book_count = 1
        for book in book_sections:
            book_count += 1
            a_tags = book.find_all('a', attrs={'class':'a-link-normal'})
            book_details_link = book_prefix['United States'] + a_tags[0]['href']
            try:
                browser.get(book_details_link)
                source = browser.page_source
                soup = BeautifulSoup(source, features='lxml')
                
                title = soup.find('span', attrs={'id':'productTitle'}).get_text().strip('\n')
                author = soup.find('a', attrs={'class':'a-link-normal contributorNameID'}).get_text()
                rating = soup.find('span', attrs={'id':'acrCustomerReviewText'}).get_text().replace(' ratings', '')
                stars = soup.find('span', attrs={'class': 'reviewCountTextLinkedHistogram noUnderline'})['title'].replace(' out of 5 stars', '')
                div = soup.find('div', attrs={'cel_widget_id':'dpx-detail-bullets_csm_instrumentation_wrapper'})
                ul_tags = div.find_all('ul', attrs={'class':'a-unordered-list a-nostyle a-vertical a-spacing-none detail-bullet-list'})
                
                details = {}
                for name in self.sub_names: details[name] = self.sub_names[name]
                if 'Title'          in self.data_fields:    details['Title']        = title
                if 'Web-Link'       in self.data_fields:    details['Web-Link']     = book_details_link
                if 'Author'         in self.data_fields:    details['Author']       = author
                if 'Ratings'        in self.data_fields:    details['Ratings']      = rating
                if 'Stars'          in self.data_fields:    details['Stars']        = stars

                li_list = ul_tags[0].find_all('li')
                for li in li_list:
                    span = li.find('span').find_all('span')
                    heading = str( span[0].get_text() ).split(':')[0].strip('\n')                        
                    data = span[1].get_text()
                    try:
                        if heading == 'Publisher':
                            data = data.split('(')
                            details[heading] = data[0]
                            details['Publication date'] = data[1].replace(')', '')
                        else:   details[heading] = data
                    except: details[heading] = data
                li = ul_tags[1].find('li')
                span = li.find('span').get_text().replace('Best Sellers Rank:', '')
                span = span.replace(' (See Top 100 in Kindle Store)', '')
                span = span.replace('\n\n\n\n','')
                span = span.replace('\n\n','\n')
                details['Best Sellers Rank'] = span
                print('K-', self.count, '. ', details, '\n')
                self.count += 1
                books.append(details)
            except: continue
            if book_count > self.book_number:   break
        
        # Saving in Excel File  
        self.write_to_excel(filename, books)
        browser.close()

    def create_excel_file(self, category_name, filename):
        # creating new excel file
        workbook = xlsxwriter.Workbook(filename)
        workbook.add_worksheet(category_name)
        workbook.close()
        workbook = op.load_workbook(filename, False)
        worksheet = workbook[category_name]
        worksheet.append(self.headers())
        workbook.save(filename)
        workbook.close()

    def write_to_excel(self, filename, books=[]):
        workbook = op.load_workbook(filename, False)
        worksheet = workbook[self.sub_names['category']]
        for book in books:
            data = []
            for data_field in self.data_fields:
                try:
                    if data_field == 'Best Sellers Rank':
                        values = book[data_field].split('\n')
                        for v in values: data.append(v)
                    else:   data.append(book[data_field])
                except: data.append('N/A')
            worksheet.append(data)
        workbook.save(filename)
        workbook.close()

    def headers(self):
        header = []
        for data_field in self.data_fields:
            header.append(data_field)
        return header


if __name__ == '__main__':
    audi = audible()
    # kind = kindle()
    audi.scrape_category()
    # kind.scrape_category()

